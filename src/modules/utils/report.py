import csv
from datetime import date, datetime, timezone
from openpyxl import load_workbook
from sqlalchemy import text
from src.modules import Database, logger
from src.modules.config import settings
import src.modules.schemas.report as schema
from src.modules.utils.utils import query_data


def generate_noon_report(imo: int, report_date: date):
    output_path = settings.OUTPUT_PATH
    template_path = settings.TEMPLATE_PATH
    daily_report = settings.DAILY_REPORT
    template_path_name = f"{template_path}/{daily_report}"

    sql = text("select v.name, :report_date from core.vessel v where v.imo = :imo;")
    cons_sql = text(
        """
select
    sum(c.consumption),
    date_trunc('day', c.date_hour) as dt
from
    core.vessel v
left outer join core.total_fuel_consumption c on v.id = c.vessel_id
where v.imo = :imo
    and date_trunc('day', c.date_hour)= :report_date
group by dt;
"""
    )
    wind_sql = text(
        """
select
    min(w.wind_speed),
    max(w.wind_speed),
    date_trunc('day', w.message_datetime) as dt
from core.vessel v
inner join core.device d on d.vessel_id = v.id
left outer join core.nmea_report_data w on d.device_id = w.device_id
where d.device_type_id = 3 and v.imo = :imo
    and date_trunc('day', w.message_datetime)= :report_date
    and w.reference = 'T' and w.wind_speed_units = 'M'
group by dt;
"""
    )
    hours_sql = text(
        """
select
    coalesce (max(c.me_running_hour) - min(c.me_running_hour), 0),
    e.type,
    date_trunc('day', c.message_datetime) as dt
from core.vessel v
inner join core.engine e on e.vessel_id = v.id
inner join core.device d on d.engine_id = e.id
left outer join core.canbus_report_data c on d.device_id = c.device_id
where d.device_type_id = 1 and v.imo = :imo
    and date_trunc('day', c.message_datetime)= :report_date
group by e.type, dt;
"""
    )

    # Get data from DB
    param = {
        "imo": imo,
        "report_date": (report_date),
    }
    try:
        engine = Database.get_engine()
        with engine.connect() as connection:
            stmt = connection.execute(sql, param)
            result = stmt.fetchone()
            cons_stmt = connection.execute(cons_sql, param)
            cons_result = cons_stmt.fetchone()
            wind_stmt = connection.execute(wind_sql, param)
            wind_result = wind_stmt.fetchone()
            hours_stmt = connection.execute(hours_sql, param)
            hours_result = hours_stmt.fetchall()
    except Exception as e:
        logger.error("DB connection error. Error: %s", e)
        raise e

    # Create report data dict
    if result:
        key_list = {
            "vessel_name": result[0],
            "report_date": result[1],
            "total_consumption": "N/A",
            "wind_speed": "N/A",
            "c_hours": 0,
            "p_hours": 0,
            "s_hours": 0,
        }
    else:
        return None, None
    if cons_result:
        key_list["total_consumption"] = cons_result[0]
    if wind_result:
        key_list["wind_speed"] = f"{wind_result[0]}-{wind_result[1]} MPS"
    if hours_result:
        for row in hours_result:
            if row[1] == "ME_CENTRE":
                key_list["c_hours"] = row[0]
            elif row[1] == "ME_PORT":
                key_list["p_hours"] = row[0]
            elif row[1] == "ME_STBD":
                key_list["s_hours"] = row[0]

    # Generate report
    file_name = f"{result[0]}_{result[1]}.xlsx"
    output_path_name = f"{output_path}/{file_name}"
    try:
        wb = load_workbook(template_path_name)
        ws = wb["Daily Report"]
        for row in ws.iter_rows(min_row=4, max_row=43, min_col=3, max_col=22):
            for cell in row:
                if cell.value in key_list:
                    cell.value = key_list[str(cell.value)]
        logger.info("output_path_name: %s, key_list: %s", output_path_name, key_list)
        wb.save(output_path_name)
    except Exception as e:
        logger.error("File to generate report: %s", e)
        raise e

    return output_path_name, file_name

def get_unified_data(
    imo: int, start_date: int, end_date: int, offset: int, limit: int | None
):
    logger.info("Getting unified data with interpolation")
    # Get data from DB with optimized RPM join with rpm
    additional_columns = ", COUNT(*) OVER () AS row_count" if limit is not None else ""
    sql = text(
        f"""
 SELECT 
    
     MAX(CASE 
        WHEN i.nr_time = c.message_datetime AND d.label_name = 'ME1' 
        THEN c.me_rpm END) AS me1_rpm,
    
    MAX(CASE 
        WHEN i.nr_time = c.message_datetime AND d.label_name = 'ME2' 
        THEN c.me_rpm END) AS me2_rpm,

    MAX(CASE 
        WHEN i.nr_time = c.message_datetime AND d.label_name = 'ME3' 
        THEN c.me_rpm END) AS me3_rpm,
    i.*
   {additional_columns}
FROM core.interpolated_data i
INNER JOIN core.vessel v ON v.id = i.vessel_id
INNER JOIN core.device d ON d.vessel_id = v.id
LEFT JOIN core.canbus_report_data c 
    ON d.device_id = c.device_id 
    AND i.nr_time = c.message_datetime -- Only match when timestamps are equa
WHERE v.imo = :imo
AND i.nr_time >= :start_date :: date and i.nr_time < :end_date :: date
GROUP BY i.nr_time, i.id, i.vessel_id 
ORDER BY i.nr_time desc
limit :limit offset :offset"""
    )
    data = query_data(
        sql,
        {
            "imo": imo,
            "start_date": datetime.fromtimestamp(start_date, tz=timezone.utc),
            "end_date": datetime.fromtimestamp(end_date, tz=timezone.utc),
            "limit": limit,
            "offset": offset,
        },
    )
    return data


#     SELECT 
#     i.*,
#     CASE 
#         WHEN i.nr_time = c.message_datetime THEN c.me_rpm
#         ELSE NULL
#     END AS me_rpm
#     {"" if limit is None else ", COUNT(*) OVER () AS row_count"}
# FROM core.interpolated_data i
# INNER JOIN core.vessel v ON v.id = i.vessel_id
# INNER JOIN core.device d ON d.vessel_id = v.id
# LEFT JOIN core.canbus_report_data c 
#     ON d.device_id = c.device_id 
#     AND i.nr_time = c.message_datetime -- Only match when timestamps are equal
# WHERE v.imo = :imo
# AND i.nr_time BETWEEN :start_date AND :end_date
# ORDER BY i.nr_time desc


# def get_unified_data(
#     imo: int, start_date: int, end_date: int, offset: int, limit: int | None
# ):
#     logger.info("Getting unified data with interpolation")
#     # Get data from DB
#     sql = text(
#         f"""select d.*{"" if limit is None else ", COUNT(*) OVER () AS row_count"}
# from interpolated_data d
# inner join vessel v on v.id = d.vessel_id
# where v.imo = :imo and d.nr_time >= :start_date :: date and d.nr_time < :end_date :: date
# order by nr_time desc
# limit :limit offset :offset"""
#     )
#     data = query_data(
#         sql,
#         {
#             "imo": imo,
#             "start_date": datetime.fromtimestamp(start_date, tz=timezone.utc),
#             "end_date": datetime.fromtimestamp(end_date, tz=timezone.utc),
#             "limit": limit,
#             "offset": offset,
#         },
#     )
#     return data


def get_unified_dataset(
    imo: int, start_date: int = 0, end_date: int = 0, offset: int = 0, limit: int = 20
) -> schema.UnifiedDatasetResponse:
    if limit < 1:
        raise ValueError("Limit must be greater than 0")
    data = get_unified_data(imo, start_date, end_date, offset, limit)
    logger.debug("Total records number found for the specified filter: %s", len(data))
    logger.info("Validate unified data")
    content = schema.UnifiedDatasetResponse(
        data=[schema.UnifiedDataset.model_validate(row) for row in data],
        page_size=limit,
        page_number=offset,
        total_pages=int(data[0][-1]/limit) if data else 0,
    )
    return content


def get_unified_file(
    imo: int,
    start_date: int = 0,
    end_date: int = 0,
    offset: int = 0,
    limit: int | None = None,
) -> str | None:
    data = get_unified_data(imo, start_date, end_date, offset, limit)
    if not data:
        logger.info("No data found for the specified filter")
        return None
    logger.info("Dump unified data to csv file")
    csv_file_path = f"{settings.TEMP_PATH}/{imo}.csv"
    with open(csv_file_path, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=data[0]._fields)
        writer.writeheader()
        for row in data:
            writer.writerow(row._asdict())
    return csv_file_path


def get_cons_summary_data(imo: int, start_date: int, end_date: int):
    logger.info("Getting consumption summary data from interpolated_data table")
    # Get data from DB
    sql = text(
        """select date_trunc('day', d.nr_time) as date,
        max(d.me1_run_hours) - min(d.me1_run_hours) as me1_run_hours,
        max(d.me2_run_hours) - min(d.me2_run_hours) as me2_run_hours,
        max(d.me3_run_hours) - min(d.me3_run_hours) as me3_run_hours,
        max(d.ae1_run_hours) - min(d.ae1_run_hours) as ae1_run_hours,
        max(d.ae2_run_hours) - min(d.ae2_run_hours) as ae2_run_hours,
        max(d.ae3_run_hours) - min(d.ae3_run_hours) as ae3_run_hours,
        max(d.ae4_run_hours) - min(d.ae4_run_hours) as ae4_run_hours,
avg(d.average_speed_gps) as avg_sog, avg(d.speed_through_water) as avg_stw,
sum(coalesce (d.me1_consumption, 0)) / 60 as sum_cons_me1,
sum(coalesce (d.me2_consumption, 0)) / 60 as sum_cons_me2,
sum(coalesce (d.me3_consumption, 0)) / 60 as sum_cons_me3,
sum(coalesce (d.me1_consumption, 0) + coalesce(d.me2_consumption, 0)
              + coalesce(d.me3_consumption, 0)) / 60 as sum_cons_me_total,
sum(coalesce (d.ae1_consumption, 0)) / 60 as sum_cons_ae1,
sum(coalesce (d.ae2_consumption, 0)) / 60 as sum_cons_ae2,
sum(coalesce (d.ae1_consumption, 0) + coalesce (d.ae2_consumption, 0)) / 60 as sum_cons_ae_total,
ceil(avg(wd.beaufourt)) as wind_scale, ceil(avg(wd.dss)) as wave_scale
from interpolated_data d
inner join vessel v on v.id = d.vessel_id
left outer join weather_data wd on wd.time = date_trunc('hour', d.nr_time)
                and wd.lat = round(d.latitude, 1) and wd.long = round(d.longitude, 1)
where v.imo = :imo and d.nr_time between :start_date :: date and :end_date :: date
group by date_trunc('day', d.nr_time)
order by date_trunc('day', d.nr_time)"""
    )
    data = query_data(
        sql,
        {
            "imo": imo,
            "start_date": datetime.fromtimestamp(start_date, tz=timezone.utc),
            "end_date": datetime.fromtimestamp(end_date, tz=timezone.utc),
        },
    )
    if not data:
        logger.info("No data found for the specified filter")
        return []
    logger.debug("Total records number found for the specified filter: %s", len(data))
    return [schema.ConsumptionSummary.model_validate(row) for row in data]


def get_client_summary_data(imo: int, start_date: int, end_date: int):
    logger.info("Getting customer report data from customer_unified_data table")
    # Get data from DB
    sql = text(
        """select date_trunc('day', d.nr_time) as date,
coalesce (d.average_speed_gps, 0) as avg_sog, coalesce(d.speed_through_water, 0) as avg_stw,
coalesce (d.me1_run_hours, 0) as sum_hours_me1,
coalesce (d.me2_run_hours, 0) as sum_hours_me2,
coalesce (d.me3_run_hours, 0) as sum_hours_me3,
coalesce (d.me1_consumption, 0) as sum_cons_me1,
coalesce (d.me2_consumption, 0) as sum_cons_me2,
coalesce (d.me3_consumption, 0) as sum_cons_me3,
coalesce (d.me1_consumption, 0) + coalesce(d.me2_consumption, 0)
              + coalesce(d.me3_consumption, 0) as sum_cons_me_total,
coalesce (d.ae1_run_hours, 0) as sum_hours_ae1,
coalesce (d.ae2_run_hours, 0) as sum_hours_ae2,
coalesce (d.ae1_consumption, 0) as sum_cons_ae1,
coalesce (d.ae2_consumption, 0) as sum_cons_ae2,
coalesce (d.ae1_consumption, 0) + coalesce (d.ae2_consumption, 0) as sum_cons_ae_total,
ceil(coalesce (wd.beaufourt, 0)) as wind_scale, ceil(coalesce (wd.dss, 0)) as wave_scale
from customer_unified_data d
inner join vessel v on v.id = d.vessel_id
left outer join weather_data wd on wd.time = date_trunc('hour', d.nr_time)
                and wd.lat = round(d.latitude, 1) and wd.long = round(d.longitude, 1)
where v.imo = :imo and d.nr_time between :start_date :: date and :end_date :: date
order by date_trunc('day', d.nr_time)"""
    )
    data = query_data(
        sql,
        {
            "imo": imo,
            "start_date": datetime.fromtimestamp(start_date, tz=timezone.utc),
            "end_date": datetime.fromtimestamp(end_date, tz=timezone.utc),
        },
    )
    if not data:
        logger.info("No data found for the specified filter")
        return []
    logger.debug("Total records number found for the specified filter: %s", len(data))
    return [schema.ConsumptionSummary.model_validate(row) for row in data]
